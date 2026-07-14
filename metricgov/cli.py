from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import yaml

PACKAGE_ROOT = Path(__file__).resolve().parent.parent
PACKS_DIR = PACKAGE_ROOT / "metric_packs"

REQUIRED_MDR_FIELDS = [
    "Status",
    "Owner",
    "Definition",
    "Source of Truth",
    "Logic / Formula",
    "Grain",
    "Time Basis",
    "Approved Use",
    "Not Approved Use",
    "Related Metrics",
    "Caveats",
    "Review Cadence",
]

EVIDENCE_FIELDS = [
    "evidence_id",
    "metric_family",
    "metric_label",
    "metric_type",
    "source_file",
    "source_type",
    "logic_found",
    "table_or_field",
    "filter_or_context",
    "date_basis",
    "owner_mentioned",
    "confidence",
]


def slugify(value: str) -> str:
    value = value.strip().lower().replace("/", " ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "metric"


def ensure_dirs(root: Path) -> None:
    for folder in ["evidence", "feedback", "artifacts", "decisions", "catalog"]:
        (root / folder).mkdir(parents=True, exist_ok=True)


def load_pack(metric_family: str) -> Dict[str, Any]:
    path = PACKS_DIR / f"{metric_family}.json"
    if not path.exists():
        available = ", ".join(p.stem for p in PACKS_DIR.glob("*.json"))
        raise SystemExit(f"Unknown metric family '{metric_family}'. Available: {available}")
    return json.loads(path.read_text(encoding="utf-8"))


def read_context(root: Path) -> Dict[str, Any]:
    path = root / "metric_context.json"
    if not path.exists():
        raise SystemExit(
            "This command must be run inside a metric governance project.\n"
            "To try the worked example: cd examples/revenue\n"
            "To start a new project: metricgov init revenue"
        )
    return json.loads(path.read_text(encoding="utf-8"))


def read_structured_decisions(path: Path) -> List[Dict[str, Any]]:
    try:
        payload = yaml.safe_load(path.read_text(encoding="utf-8"))
    except yaml.YAMLError as exc:
        raise SystemExit(f"Invalid workshop decisions YAML: {exc}") from exc
    if payload is None:
        return []
    decisions = payload.get("decisions") if isinstance(payload, dict) else payload
    if not isinstance(decisions, list) or not all(isinstance(item, dict) for item in decisions):
        raise SystemExit("Workshop decisions YAML must contain a 'decisions' list.")
    for index, decision in enumerate(decisions, 1):
        if not decision.get("metric"):
            raise SystemExit(f"Workshop decision {index} is missing 'metric'.")
        if "owner_confirmed" in decision and not isinstance(decision["owner_confirmed"], bool):
            raise SystemExit(f"Workshop decision {index} 'owner_confirmed' must be true or false.")
    return decisions


def decision_value(decision: Dict[str, Any], key: str) -> Any:
    raw = decision.get(key)
    if isinstance(raw, dict):
        if decision.get("owner_confirmed") is True or raw.get("proposed") is True:
            return raw.get("value")
        return None
    return raw if decision.get("owner_confirmed") is True else None


def display_decision_value(value: Any, *, bullets: bool = False) -> str:
    if isinstance(value, list):
        return "\n".join(f"- {item}" for item in value) if bullets else ", ".join(str(item) for item in value)
    return str(value).strip() if value is not None else ""


def find_decision(decisions: List[Dict[str, Any]], metric_name: str, metric_type: str) -> Dict[str, Any]:
    targets = {slugify(metric_name), slugify(metric_type)}
    return next((item for item in decisions if slugify(str(item.get("metric", ""))) in targets), {})


def write_context(root: Path, metric_family: str) -> None:
    pack = load_pack(metric_family)
    context = {
        "metric_family": metric_family,
        "display_name": pack.get("display_name", metric_family),
        "business_problem": "Metric definitions are inconsistent or unclear across teams and reports.",
        "primary_decision": "To be confirmed with stakeholders.",
        "teams_involved": pack.get("typical_owners", []),
        "status": "discovery",
    }
    (root / "metric_context.json").write_text(json.dumps(context, indent=2), encoding="utf-8")


def normalise_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def contains_term(text: str, term: str) -> bool:
    # Normalise underscores, hyphens, punctuation, and SQL aliases so pipeline_revenue matches pipeline revenue.
    text_n = f" {normalise_text(text)} "
    term_n = f" {normalise_text(term)} "
    return term_n in text_n


def match_terms(text: str, pack: Dict[str, Any]) -> List[Tuple[str, str]]:
    matches: List[Tuple[str, str]] = []
    for metric_type, terms in pack.get("terms", {}).items():
        for term in terms:
            if contains_term(text, term):
                matches.append((term, metric_type))
    # Also match family display name, e.g. Revenue or Churn, as ambiguous generic label.
    display = pack.get("display_name", pack["metric_family"])
    if contains_term(text, display):
        matches.append((display, "ambiguous_generic"))
    # de-dupe preserving order
    seen = set()
    deduped = []
    for label, mtype in matches:
        key = (label.lower(), mtype)
        if key not in seen:
            deduped.append((label, mtype))
            seen.add(key)
    return deduped


def maybe_owner(text: str, pack: Dict[str, Any]) -> str:
    text_l = text.lower()
    found = []
    for owner in pack.get("typical_owners", []):
        if owner.lower() in text_l:
            found.append(owner)
    return "; ".join(found)


def extract_sql_info(text: str) -> Dict[str, str]:
    table = ""
    filters = ""
    aliases = []
    for m in re.finditer(r"\bas\s+([a-zA-Z_][\w]*)", text, flags=re.I):
        aliases.append(m.group(1))
    tables = re.findall(r"\bfrom\s+([a-zA-Z_][\w\.]*|`[^`]+`|\"[^\"]+\")", text, flags=re.I)
    joins = re.findall(r"\bjoin\s+([a-zA-Z_][\w\.]*|`[^`]+`|\"[^\"]+\")", text, flags=re.I)
    if tables or joins:
        table = "; ".join(tables + joins)
    where = re.search(r"\bwhere\b(.+?)(\bgroup\s+by\b|\border\s+by\b|\blimit\b|$)", text, flags=re.I | re.S)
    if where:
        filters = re.sub(r"\s+", " ", where.group(1)).strip()[:500]
    return {"aliases": "; ".join(aliases), "table": table, "filters": filters}


def scan_text_file(path: Path, pack: Dict[str, Any]) -> List[Dict[str, str]]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    matches = match_terms(text, pack)
    rows = []
    sql_info = extract_sql_info(text) if path.suffix.lower() == ".sql" else {"aliases": "", "table": "", "filters": ""}
    for label, metric_type in matches:
        # Capture a short context around first occurrence.
        idx = text.lower().find(label.lower())
        context = ""
        if idx >= 0:
            context = re.sub(r"\s+", " ", text[max(0, idx - 120): idx + 240]).strip()
        rows.append({
            "metric_family": pack["metric_family"],
            "metric_label": label,
            "metric_type": metric_type,
            "source_file": str(path),
            "source_type": path.suffix.lower().lstrip("."),
            "logic_found": sql_info.get("aliases") or context,
            "table_or_field": sql_info.get("table", ""),
            "filter_or_context": sql_info.get("filters") or context,
            "date_basis": infer_date_basis(text),
            "owner_mentioned": maybe_owner(text, pack),
            "confidence": "medium" if sql_info.get("table") or sql_info.get("filters") else "low",
        })
    return rows


def infer_date_basis(text: str) -> str:
    text_l = text.lower()
    candidates = [
        "close date", "invoice date", "recognition month", "recognised month", "recognized month",
        "created date", "cancellation date", "service end date", "month", "period", "date"
    ]
    return "; ".join(c for c in candidates if c in text_l)[:200]


def scan_csv_file(path: Path, pack: Dict[str, Any]) -> List[Dict[str, str]]:
    rows = []
    try:
        with path.open(newline="", encoding="utf-8-sig", errors="ignore") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
    except Exception as exc:
        return []
    header_text = " | ".join(headers)
    matches = match_terms(header_text, pack)
    for label, metric_type in matches:
        matching_headers = [h for h in headers if label.lower() in h.lower() or any(t.lower() in h.lower() for t in pack.get("terms", {}).get(metric_type, []))]
        rows.append({
            "metric_family": pack["metric_family"],
            "metric_label": label,
            "metric_type": metric_type,
            "source_file": str(path),
            "source_type": "csv",
            "logic_found": "Detected from CSV header",
            "table_or_field": "; ".join(matching_headers) or header_text[:300],
            "filter_or_context": header_text[:500],
            "date_basis": infer_date_basis(header_text),
            "owner_mentioned": maybe_owner(header_text, pack),
            "confidence": "low",
        })
    return rows


def scan_xlsx_file(path: Path, pack: Dict[str, Any]) -> List[Dict[str, str]]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return []
    rows = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return []
    for ws in wb.worksheets:
        values = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
            values.extend([str(v) for v in row if v is not None])
        text = f"Sheet: {ws.title}\n" + " | ".join(values)
        matches = match_terms(text, pack)
        for label, metric_type in matches:
            rows.append({
                "metric_family": pack["metric_family"],
                "metric_label": label,
                "metric_type": metric_type,
                "source_file": str(path),
                "source_type": "xlsx",
                "logic_found": f"Detected from workbook sheet '{ws.title}' headers/first rows",
                "table_or_field": " | ".join(values[:20])[:500],
                "filter_or_context": text[:500],
                "date_basis": infer_date_basis(text),
                "owner_mentioned": maybe_owner(text, pack),
                "confidence": "low",
            })
    return rows


def write_evidence_log(root: Path, rows: List[Dict[str, str]]) -> Path:
    out = root / "artifacts" / "01_evidence_log.csv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=EVIDENCE_FIELDS)
        writer.writeheader()
        for idx, row in enumerate(rows, 1):
            row = {k: row.get(k, "") for k in EVIDENCE_FIELDS}
            row["evidence_id"] = row.get("evidence_id") or f"E{idx:03d}"
            writer.writerow(row)
    return out


def read_evidence(root: Path) -> List[Dict[str, str]]:
    path = root / "artifacts" / "01_evidence_log.csv"
    if not path.exists():
        raise SystemExit("Evidence log not found. Run `metricgov scan` first.")
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def command_init(args: argparse.Namespace) -> None:
    root = Path.cwd()
    ensure_dirs(root)
    write_context(root, args.metric_family)
    (root / "evidence" / "README.md").write_text(
        "# Evidence folder\n\nAdd SQL, CSV, Markdown, text, or Excel evidence files here. Prefer metadata, sample headers, SQL, and notes over sensitive raw data.\n",
        encoding="utf-8",
    )
    (root / "feedback" / "workshop_notes.md").write_text(
        "# Workshop notes\n\nAfter stakeholder alignment, record confirmed decisions here.\n",
        encoding="utf-8",
    )
    print(f"Initialized metric governance project for '{args.metric_family}'.")


def command_scan(args: argparse.Namespace) -> None:
    root = Path.cwd()
    context = read_context(root)
    pack = load_pack(context["metric_family"])
    evidence_dir = root / "evidence"
    if not evidence_dir.exists():
        raise SystemExit("No evidence/ folder found. Run `metricgov init <metric_family>` first.")
    rows: List[Dict[str, str]] = []
    for path in sorted(evidence_dir.rglob("*")):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix in [".md", ".txt", ".sql"]:
            rows.extend(scan_text_file(path, pack))
        elif suffix == ".csv":
            rows.extend(scan_csv_file(path, pack))
        elif suffix in [".xlsx", ".xlsm"]:
            rows.extend(scan_xlsx_file(path, pack))
    out = write_evidence_log(root, rows)
    print(f"Scanned evidence. Found {len(rows)} metric evidence rows. Wrote {out.relative_to(root)}")


def group_by(rows: Iterable[Dict[str, str]], key: str) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        grouped.setdefault(row.get(key, "unknown") or "unknown", []).append(row)
    return grouped


def command_classify(args: argparse.Namespace) -> None:
    root = Path.cwd()
    context = read_context(root)
    pack = load_pack(context["metric_family"])
    rows = read_evidence(root)
    family_map = root / "artifacts" / "02_metric_family_map.md"
    ambiguity = root / "artifacts" / "03_ambiguity_register.md"

    by_type = group_by(rows, "metric_type")
    with family_map.open("w", encoding="utf-8") as f:
        f.write(f"# Metric Family Map: {pack.get('display_name')}\n\n")
        f.write("This file groups discovered metric evidence into metric-family candidates. It does not certify definitions.\n\n")
        for mtype, items in sorted(by_type.items()):
            f.write(f"## {mtype}\n\n")
            labels = sorted({i.get("metric_label", "") for i in items if i.get("metric_label")})
            sources = sorted({Path(i.get("source_file", "")).name for i in items if i.get("source_file")})
            f.write(f"- Labels found: {', '.join(labels) or 'None'}\n")
            f.write(f"- Evidence count: {len(items)}\n")
            f.write(f"- Source files: {', '.join(sources) or 'None'}\n")
            f.write("- Status: needs stakeholder confirmation\n\n")

    issues = build_ambiguity_issues(rows, pack)
    with ambiguity.open("w", encoding="utf-8") as f:
        f.write(f"# Ambiguity Register: {pack.get('display_name')}\n\n")
        if not issues:
            f.write("No obvious ambiguity issues were detected from the supplied evidence. This does not mean the metric is certified.\n")
        else:
            f.write("| Issue ID | Type | Evidence | Risk | Suggested next step |\n")
            f.write("|---|---|---|---|---|\n")
            for idx, issue in enumerate(issues, 1):
                f.write(f"| I{idx:03d} | {issue['type']} | {issue['evidence']} | {issue['risk']} | {issue['next_step']} |\n")
    print(f"Classified metric evidence. Wrote {family_map.relative_to(root)} and {ambiguity.relative_to(root)}")


def build_ambiguity_issues(rows: List[Dict[str, str]], pack: Dict[str, Any]) -> List[Dict[str, str]]:
    issues: List[Dict[str, str]] = []
    display = pack.get("display_name", pack["metric_family"])
    generic_rows = [r for r in rows if r.get("metric_type") == "ambiguous_generic" or r.get("metric_label", "").lower() == display.lower()]
    if generic_rows:
        issues.append({
            "type": "generic_label_without_certification",
            "evidence": f"Generic label '{display}' appears in {len(generic_rows)} evidence rows.",
            "risk": "The same label may hide multiple business meanings.",
            "next_step": "Ask owners whether this should be split into more specific certified metrics.",
        })
    by_label = group_by(rows, "metric_label")
    for label, items in by_label.items():
        sources = {Path(i.get("source_file", "")).name for i in items}
        types = {i.get("metric_type", "") for i in items}
        if len(sources) > 1 and len(types) > 0:
            issues.append({
                "type": "same_label_multiple_sources",
                "evidence": f"'{label}' appears in multiple files: {', '.join(sorted(sources))}.",
                "risk": "Different reports may use the same label with different logic or ownership.",
                "next_step": "Confirm source of truth, owner, and approved use for this label.",
            })
    by_type = group_by(rows, "metric_type")
    non_generic_types = [t for t in by_type.keys() if t not in ["ambiguous_generic", "unknown"]]
    if len(non_generic_types) > 1:
        issues.append({
            "type": "metric_family_has_multiple_variants",
            "evidence": f"Detected variants: {', '.join(sorted(non_generic_types))}.",
            "risk": "Related metrics may be incorrectly merged or compared without caveats.",
            "next_step": "Create distinct Metric Decision Records for legitimate variants.",
        })
    for r in rows:
        if not r.get("owner_mentioned"):
            issues.append({
                "type": "owner_gap",
                "evidence": f"{r.get('metric_label')} in {Path(r.get('source_file','')).name} has no owner mentioned.",
                "risk": "Metric cannot be certified without an accountable owner.",
                "next_step": "Identify the business owner before publishing as certified.",
            })
            break
    return issues[:20]


def command_workshop(args: argparse.Namespace) -> None:
    root = Path.cwd()
    context = read_context(root)
    pack = load_pack(context["metric_family"])
    ambiguity_path = root / "artifacts" / "03_ambiguity_register.md"
    if not ambiguity_path.exists():
        raise SystemExit("Ambiguity register not found. Run `metricgov classify` first.")
    out = root / "artifacts" / "04_workshop_pack.md"
    ambiguity_text = ambiguity_path.read_text(encoding="utf-8")
    with out.open("w", encoding="utf-8") as f:
        f.write(f"# Stakeholder Alignment Workshop Pack: {pack.get('display_name')}\n\n")
        f.write("## Purpose\n\n")
        f.write("Align business stakeholders on what each metric variant means, who owns it, where it should be used, and what should be certified or deprecated.\n\n")
        f.write("## Pre-read summary\n\n")
        f.write("The scan identified metric evidence that requires human confirmation. The goal of the workshop is not to force one definition, but to decide which variants are legitimate, how they should be named, and which uses are approved.\n\n")
        f.write("## Current ambiguity register\n\n")
        f.write(ambiguity_text + "\n\n")
        f.write("## Stakeholder questions\n\n")
        for owner, questions in pack.get("owner_questions", {}).items():
            f.write(f"### {owner}\n\n")
            for q in questions:
                f.write(f"- {q}\n")
            f.write("\n")
        f.write("## Recommended agenda\n\n")
        agenda = [
            "Confirm the decision context for this metric family.",
            "Review discovered metric variants and evidence.",
            "Classify each variant: certify, rename, split, deprecate, or investigate.",
            "Assign business owner and source of truth for each certified metric.",
            "Confirm approved use and not-approved use.",
            "Agree open questions, deadlines, and next review cadence.",
        ]
        for idx, item in enumerate(agenda, 1):
            f.write(f"{idx}. {item}\n")
        f.write("\n## Decision options\n\n")
        f.write("- **Certify:** the metric has owner-confirmed definition, source, logic, grain, time basis, usage, and caveats.\n")
        f.write("- **Rename:** the current label is ambiguous and should be changed in reports.\n")
        f.write("- **Split:** one generic label should become multiple legitimate certified metrics.\n")
        f.write("- **Deprecate:** a label or report should no longer be used for decision-making.\n")
        f.write("- **Investigate:** evidence is insufficient for a decision.\n")
    print(f"Generated workshop pack: {out.relative_to(root)}")


def command_record(args: argparse.Namespace) -> None:
    root = Path.cwd()
    ensure_dirs(root)
    context = read_context(root)
    pack = load_pack(context["metric_family"])
    feedback_path = Path(args.from_file) if args.from_file else root / "feedback" / "workshop_notes.md"
    if not feedback_path.is_absolute():
        feedback_path = root / feedback_path
    feedback = feedback_path.read_text(encoding="utf-8", errors="ignore") if feedback_path.exists() else ""
    structured = feedback_path.suffix.lower() in [".yaml", ".yml"]
    decisions = read_structured_decisions(feedback_path) if structured and feedback_path.exists() else []
    rows = read_evidence(root)
    by_type = {k: v for k, v in group_by(rows, "metric_type").items() if k != "ambiguous_generic"}
    if not by_type:
        by_type = {pack["metric_family"]: rows}
    decision_log = root / "artifacts" / "05_decision_log.md"
    decision_log.parent.mkdir(parents=True, exist_ok=True)
    created = []
    for idx, (metric_type, items) in enumerate(sorted(by_type.items()), 1):
        labels = sorted({i.get("metric_label", "") for i in items if i.get("metric_label")})
        metric_name = choose_metric_name(metric_type, labels, pack)
        decision = find_decision(decisions, metric_name, metric_type)
        owner = infer_owner_for_metric(metric_type, "" if structured else feedback, pack) or (pack.get("typical_owners") or ["TBD"])[0]
        path = root / "decisions" / f"MDR-{idx:03d}-{slugify(metric_name)}.md"
        path.write_text(render_mdr(metric_name, metric_type, owner, labels, items, feedback, pack, decision), encoding="utf-8")
        created.append(path)
    with decision_log.open("w", encoding="utf-8") as f:
        f.write(f"# Decision Log: {pack.get('display_name')}\n\n")
        if structured:
            f.write("Generated Metric Decision Records from structured workshop decisions. Only explicitly proposed or owner-confirmed YAML values populate structured fields.\n\n")
        else:
            f.write("Generated draft Metric Decision Records from evidence and workshop feedback. Review with owners before marking as Approved.\n\n")
        if feedback:
            f.write("## Workshop feedback used\n\n")
            f.write(feedback.strip() + "\n\n")
        f.write("## Created MDRs\n\n")
        for p in created:
            f.write(f"- {p.relative_to(root)}\n")
    print(f"Recorded decisions. Created {len(created)} MDR draft(s).")


def choose_metric_name(metric_type: str, labels: List[str], pack: Dict[str, Any]) -> str:
    # Prefer the governed metric type name over a noisy evidence label such as "closed won" or "credits".
    if metric_type and metric_type not in ["ambiguous_generic", "unknown"]:
        name = metric_type.replace("_", " ").title()
        return name.replace("Arr", "ARR").replace("Mrr", "MRR").replace("Nrr", "NRR")
    if labels:
        display = pack.get("display_name", "").lower()
        explicit = [l for l in labels if l.lower() != display]
        label = sorted(explicit or labels, key=len)[0]
        return label.title().replace("Arr", "ARR").replace("Mrr", "MRR").replace("Nrr", "NRR")
    return pack.get("display_name", "Metric")



def infer_owner_for_metric(metric_type: str, feedback: str, pack: Dict[str, Any]) -> str:
    # Metric-specific owner hints keep generated drafts closer to practical governance.
    hints = {
        "booked_revenue": "Sales Ops",
        "pipeline_value": "Marketing",
        "recognised_revenue": "Finance",
        "net_revenue": "Finance",
        "gross_revenue": "Finance",
        "logo_churn": "Customer Success",
        "revenue_churn": "Finance",
        "subscription_churn": "Customer Success",
        "active_customer": "Product",
        "monthly_active": "Product",
        "engaged_customer": "Product",
        "paying_customer": "Finance",
        "transacting_customer": "Analytics",
    }
    if metric_type in hints:
        return hints[metric_type]
    return infer_owner_from_feedback(feedback, pack)

def infer_owner_from_feedback(feedback: str, pack: Dict[str, Any]) -> str:
    lower = feedback.lower()
    for owner in pack.get("typical_owners", []):
        if owner.lower() in lower:
            return owner
    return ""


def render_mdr(metric_name: str, metric_type: str, owner: str, labels: List[str], items: List[Dict[str, str]], feedback: str, pack: Dict[str, Any], decision: Optional[Dict[str, Any]] = None) -> str:
    decision = decision or {}
    source_files = sorted({Path(i.get("source_file", "")).name for i in items if i.get("source_file")})
    logic = next((i.get("logic_found", "") for i in items if i.get("logic_found")), "TBD")
    source = next((i.get("table_or_field", "") for i in items if i.get("table_or_field")), "TBD")
    date_basis = next((i.get("date_basis", "") for i in items if i.get("date_basis")), "TBD")
    values = {
        "Status": "Proposed",
        "Owner": owner or "TBD",
        "Definition": "TBD — requires owner confirmation. This draft was generated from evidence and workshop notes.",
        "Source of Truth": source or "TBD",
        "Logic / Formula": logic or "TBD",
        "Grain": "TBD",
        "Time Basis": date_basis or "TBD",
        "Approved Use": "TBD — confirm where this metric is allowed to be used.",
        "Not Approved Use": "TBD — confirm where this metric should not be used.",
        "Related Metrics": ", ".join(labels) if labels else pack.get("display_name", ""),
        "Caveats": f"- Generated draft. Do not mark as certified until owner confirms definition, source, grain, time basis, and usage boundaries.\n- Evidence files reviewed: {', '.join(source_files) if source_files else 'None'}.",
        "Review Cadence": "TBD",
    }
    yaml_fields = {
        "status": "Status", "owner": "Owner", "definition": "Definition",
        "source_of_truth": "Source of Truth", "logic_formula": "Logic / Formula",
        "grain": "Grain", "time_basis": "Time Basis", "approved_use": "Approved Use",
        "not_approved_use": "Not Approved Use", "related_metrics": "Related Metrics",
        "caveats": "Caveats", "review_cadence": "Review Cadence",
    }
    for yaml_key, mdr_field in yaml_fields.items():
        trusted = decision_value(decision, yaml_key)
        if trusted is not None:
            values[mdr_field] = display_decision_value(trusted, bullets=yaml_key == "caveats")
    owner_confirmed = decision.get("owner_confirmed") is True
    if values["Status"].strip().lower() == "certified":
        missing = [
            field for field in REQUIRED_MDR_FIELDS
            if not values[field].strip() or values[field].strip().upper() == "TBD" or values[field].strip().startswith("TBD")
        ]
        if not owner_confirmed or missing:
            reason = "owner_confirmed must be true" if not owner_confirmed else f"missing required fields: {', '.join(missing)}"
            raise SystemExit(f"Cannot certify {metric_name}: {reason}.")
    naming_decision = display_decision_value(decision_value(decision, "naming_decision"))
    owner_confirmation_section = f"\n## Owner Confirmed\n{'true' if owner_confirmed else 'false'}\n" if decision else ""
    naming_section = f"\n## Naming Decision\n{naming_decision or 'None recorded.'}\n" if decision else ""
    return f"""# MDR: {metric_name}

## Status
{values['Status']}

## Owner
{values['Owner']}
{owner_confirmation_section}

## Definition
{values['Definition']}

## Source of Truth
{values['Source of Truth']}

## Logic / Formula
{values['Logic / Formula']}

## Grain
{values['Grain']}

## Time Basis
{values['Time Basis']}

## Approved Use
{values['Approved Use']}

## Not Approved Use
{values['Not Approved Use']}
{naming_section}

## Related Metrics
{values['Related Metrics']}

## Caveats
{values['Caveats']}

## Open Questions
- What is the exact business definition?
- What source system or table is the source of truth?
- What formula, filters, exclusions, and time basis should be used?
- What decisions is this metric approved for?
- Which similar labels should be renamed, split, or deprecated?

## Review Cadence
{values['Review Cadence']}

## Change History
- Draft generated by Metric Governance Agent.

## Workshop Feedback Snapshot
{feedback.strip() if feedback.strip() else 'No workshop feedback supplied.'}
"""


def parse_mdr(path: Path) -> Dict[str, str]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    result = {"Metric": path.stem}
    title = re.search(r"^#\s+MDR:\s+(.+)$", text, flags=re.M)
    if title:
        result["Metric"] = title.group(1).strip()
    for field in REQUIRED_MDR_FIELDS:
        pattern = rf"^##\s+{re.escape(field)}\s*$\n(.*?)(?=^##\s+|\Z)"
        m = re.search(pattern, text, flags=re.M | re.S)
        result[field] = m.group(1).strip() if m else ""
    return result


def command_publish(args: argparse.Namespace) -> None:
    root = Path.cwd()
    ensure_dirs(root)
    mdrs = sorted((root / "decisions").glob("MDR-*.md"))
    if not mdrs:
        raise SystemExit("No MDR files found. Run `metricgov record` first.")
    catalog_dir = root / "catalog"
    catalog_dir.mkdir(parents=True, exist_ok=True)
    rows = [parse_mdr(p) for p in mdrs]
    csv_path = catalog_dir / "business_metric_catalog.csv"
    fields = ["Metric"] + REQUIRED_MDR_FIELDS
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})
    md_path = catalog_dir / "business_metric_dictionary.md"
    with md_path.open("w", encoding="utf-8") as f:
        f.write("# Business Metric Dictionary\n\n")
        f.write("Generated from Metric Decision Records. Draft or incomplete metrics should not be treated as certified.\n\n")
        for row in rows:
            f.write(f"## {row.get('Metric')}\n\n")
            f.write(f"- Status: {row.get('Status', '')}\n")
            f.write(f"- Owner: {row.get('Owner', '')}\n")
            f.write(f"- Definition: {one_line(row.get('Definition', ''))}\n")
            f.write(f"- Source of truth: {one_line(row.get('Source of Truth', ''))}\n")
            f.write(f"- Approved use: {one_line(row.get('Approved Use', ''))}\n")
            f.write(f"- Not approved use: {one_line(row.get('Not Approved Use', ''))}\n\n")
    print(f"Published catalog: {csv_path.relative_to(root)} and {md_path.relative_to(root)}")


def one_line(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip())[:300]


def read_markdown_section(text: str, heading: str) -> str:
    pattern = rf"^##\s+{re.escape(heading)}\s*$\n(.*?)(?=^##\s+|\Z)"
    match = re.search(pattern, text, flags=re.M | re.S)
    return match.group(1).strip() if match else ""


def parse_naming_decision(value: str) -> Tuple[str, str, bool]:
    if not value or value.strip().lower() == "none recorded.":
        return "", "", False
    quoted = re.search(
        r"\brename\s+[\"'“‘](.+?)[\"'”’]\s+to\s+[\"'“‘](.+?)[\"'”’]",
        value,
        flags=re.I | re.S,
    )
    if quoted:
        return one_line(quoted.group(1)), one_line(quoted.group(2)), True
    plain = re.search(r"\brename\s+(.+?)\s+to\s+(.+?)(?:[.\n]|$)", value, flags=re.I | re.S)
    if plain:
        current = one_line(plain.group(1)).strip(" \"'“”‘’")
        recommended = one_line(plain.group(2)).strip(" \"'“”‘’")
        return current, recommended, bool(current and recommended)
    return "", "", False


def markdown_cell(value: str) -> str:
    return one_line(value).replace("|", "\\|") or "—"


def command_change_plan(args: argparse.Namespace) -> None:
    root = Path.cwd()
    ensure_dirs(root)
    evidence = read_evidence(root)
    catalog_by_metric: Dict[str, Dict[str, str]] = {}
    catalog_path = root / "catalog" / "business_metric_catalog.csv"
    if catalog_path.exists():
        with catalog_path.open(newline="", encoding="utf-8-sig") as f:
            catalog_by_metric = {row.get("Metric", ""): row for row in csv.DictReader(f)}

    changes: List[Dict[str, str]] = []
    naming_count = 0
    for path in sorted((root / "decisions").glob("MDR-*.md")):
        text = path.read_text(encoding="utf-8", errors="ignore")
        mdr = parse_mdr(path)
        naming = read_markdown_section(text, "Naming Decision")
        if not naming or naming.lower() == "none recorded.":
            continue
        naming_count += 1
        current, recommended, parsed = parse_naming_decision(naming)
        confirmed = read_markdown_section(text, "Owner Confirmed").lower() == "true"
        catalog_row = catalog_by_metric.get(mdr.get("Metric", ""), {})
        owner = catalog_row.get("Owner") or mdr.get("Owner", "") or "TBD"
        if not parsed:
            changes.append({
                "current": naming,
                "recommended": "Review required",
                "evidence": "Not matched",
                "source_type": "",
                "reason": "Naming decision could not be parsed confidently",
                "owner": owner,
                "status": "Review required",
            })
            continue
        matches = []
        for row in evidence:
            searchable = " ".join([
                row.get("metric_label", ""), row.get("logic_found", ""),
                row.get("filter_or_context", ""), row.get("table_or_field", ""),
            ])
            if contains_term(searchable, current):
                matches.append(row)
        if not matches:
            changes.append({
                "current": current,
                "recommended": recommended,
                "evidence": "Not matched",
                "source_type": "",
                "reason": "Naming decision found; affected evidence needs review",
                "owner": owner,
                "status": "Review required",
            })
            continue
        seen = set()
        for row in matches:
            source_file = row.get("source_file", "")
            source_name = source_file.replace("\\", "/").rsplit("/", 1)[-1] or "Unknown"
            key = (source_name, row.get("source_type", ""))
            if key in seen:
                continue
            seen.add(key)
            changes.append({
                "current": current,
                "recommended": recommended,
                "evidence": source_name,
                "source_type": row.get("source_type", "") or "Unknown",
                "reason": "Owner-confirmed naming decision" if confirmed else "Naming decision requires owner confirmation",
                "owner": owner,
                "status": "Ready to rename" if confirmed else "Review required",
            })

    out = root / "artifacts" / "07_dashboard_change_plan.md"
    ready = sum(item["status"] == "Ready to rename" for item in changes)
    review = sum(item["status"] == "Review required" for item in changes)
    with out.open("w", encoding="utf-8") as f:
        f.write("# Dashboard Change Plan\n\n")
        f.write("This plan translates recorded metric naming decisions into candidate report, dashboard, SQL, and evidence changes. Review each item before implementation.\n\n")
        f.write("## Summary\n\n")
        f.write(f"- Naming decisions found: {naming_count}\n")
        f.write(f"- Ready to rename: {ready}\n")
        f.write(f"- Review required: {review}\n")
        f.write(f"- No action: {1 if naming_count == 0 else 0}\n\n")
        f.write("## Change plan\n\n")
        f.write("| Current Label | Recommended Label | Affected Evidence | Source Type | Reason | Owner | Status |\n")
        f.write("|---|---|---|---|---|---|---|\n")
        if not changes:
            f.write("| — | — | — | — | No naming decisions found. | — | No action |\n")
        else:
            for item in changes:
                fields = [item[key] for key in ["current", "recommended", "evidence", "source_type", "reason", "owner", "status"]]
                f.write("| " + " | ".join(markdown_cell(value) for value in fields) + " |\n")
        f.write("\n## Notes / limitations\n\n")
        f.write("- This plan does not modify dashboards, reports, SQL, or evidence files.\n")
        f.write("- Matches use labels and captured evidence text; indirect or dynamic references may be missed.\n")
        f.write("- Unclear naming decisions or unmatched evidence remain Review required.\n")
    print(f"Generated dashboard change plan: {out.relative_to(root)}. Ready: {ready}. Review required: {review}.")


def command_check(args: argparse.Namespace) -> None:
    root = Path.cwd()
    ensure_dirs(root)
    report = root / "artifacts" / "06_governance_check.md"
    mdrs = sorted((root / "decisions").glob("MDR-*.md"))
    failures = 0
    with report.open("w", encoding="utf-8") as f:
        f.write("# Governance Check Report\n\n")
        if not mdrs:
            f.write("No MDR files found.\n")
            failures += 1
        for path in mdrs:
            data = parse_mdr(path)
            f.write(f"## {path.name}\n\n")
            missing = []
            for field in REQUIRED_MDR_FIELDS:
                value = data.get(field, "").strip()
                if not value or value.upper() == "TBD" or value.startswith("TBD"):
                    missing.append(field)
            if missing:
                failures += 1
                f.write("Status: FAIL\n\n")
                for field in missing:
                    f.write(f"- Missing or incomplete: {field}\n")
            else:
                f.write("Status: PASS\n")
            f.write("\n")
    print(f"Governance check complete. Failures: {failures}. Report: {report.relative_to(root)}")
    if args.fail_on_error and failures:
        raise SystemExit(1)


def command_prepare(args: argparse.Namespace) -> None:
    command_scan(args)
    command_classify(args)
    command_workshop(args)


def command_finalize(args: argparse.Namespace) -> None:
    command_record(args)
    command_publish(args)
    command_change_plan(args)
    command_check(args)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="metricgov", description="Metric Governance Agent CLI")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("init", help="Initialize a metric governance project")
    p.add_argument("metric_family", choices=[p.stem for p in PACKS_DIR.glob("*.json")])
    p.set_defaults(func=command_init)

    p = sub.add_parser("scan", help="Scan evidence files and create evidence log")
    p.set_defaults(func=command_scan)

    p = sub.add_parser("classify", help="Classify metric ambiguity")
    p.set_defaults(func=command_classify)

    p = sub.add_parser("workshop", help="Generate stakeholder workshop pack")
    p.set_defaults(func=command_workshop)

    p = sub.add_parser("record", help="Generate draft Metric Decision Records from feedback")
    p.add_argument("--from", dest="from_file", default=None, help="Path to workshop feedback notes")
    p.set_defaults(func=command_record)

    p = sub.add_parser("publish", help="Publish business metric catalog from MDRs")
    p.set_defaults(func=command_publish)

    p = sub.add_parser("change-plan", help="Generate dashboard and report label change plan")
    p.set_defaults(func=command_change_plan)

    p = sub.add_parser("check", help="Validate MDR governance completeness")
    p.add_argument("--fail-on-error", action="store_true", help="Return non-zero exit code if checks fail")
    p.set_defaults(func=command_check)

    p = sub.add_parser("prepare", help="Run scan -> classify -> workshop")
    p.set_defaults(func=command_prepare)

    p = sub.add_parser("finalize", help="Run record -> publish -> change-plan -> check")
    p.add_argument("--from", dest="from_file", default=None, help="Path to workshop feedback notes")
    p.add_argument("--fail-on-error", action="store_true", help="Return non-zero exit code if checks fail")
    p.set_defaults(func=command_finalize)

    return parser


def main(argv: Optional[List[str]] = None) -> None:
    parser = build_parser()
    args = parser.parse_args(argv)
    args.func(args)


if __name__ == "__main__":
    main()
